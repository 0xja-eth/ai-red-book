import os
import shutil
import time
from abc import abstractmethod
from dataclasses import dataclass
from enum import Enum

import openpyxl
from dataclasses_json import dataclass_json
from openpyxl.styles import Font, Alignment

from src.config import config_loader
from src.config.config_loader import get
from src.core.state_manager import initial_state, get_state, set_state
from src.utils import openai_utils, api_utils

INPUT_ROOT = config_loader.file("./input")
OUTPUT_ROOT = config_loader.file("./output")

TITLE_PROMPT_FILE = "./title_prompt.txt"
CONTENT_PROMPT_FILE = "./content_prompt.txt"

TEMPLATE_FILE = "./templates.xlsx"

COL_WIDTH = {
    "ID": 12,
    "标题": 32,
    "内容": 128,
    "素材": 16,
    "标题提示词": 32,
    "内容提示词": 32,
    "创建时间": 16,
    "更新时间": 16
}

# DEFAULT_PROMPTS = {
#     "title": """假设你是一个小红书运营高手，我们要在小红书上发布笔记来推广一款产品。产品名：小黑子周边，产品描述：小黑子指的是蔡徐坤，因为蔡徐坤喜欢中分，穿吊带，唱鸡你太美，打篮球，练习时长两年半，所以被称为小黑子，并因此出现了许多周边。注意，示例里的产品和我们产品不一致，不要照搬示例！。我希望你能给我一个不超过20字的小红书笔记标题，标题的形式可以参考以下示例：
#
# 1. 🐶宠物用品一站式批发&一件代发
#
# 2. 源头工厂|一站式宠物用品打造🐾
#
# 3. 🐾一站式宠物用品批发代发，源头
#
# 注意，只需要给出一个标题，且标题里不要出现编号。""",
#     "content": """假设你是一个小红书运营高手，我们要在小红书上发布笔记来推广一款产品。产品名：小黑子周边，产品描述：小黑子指的是蔡徐坤，因为蔡徐坤喜欢中分，穿吊带，唱鸡你太美，打篮球，练习时长两年半，所以被称为小黑子，并因此出现了许多周边。注意，示例里的产品和我们产品不一致，不要照搬示例！，笔记的标题为 %s。我希望你能基于此标题写一篇小红书笔记正文。正文的形式可以参考以下示例：
#
# 1. 🐕‍🦺 中高端原创宠物用品供应商！
#
# 🌎 关于海外市场
#
# 我们的业务在海外市场沉淀几年，现已覆盖多个国家！
#
# 🤝 合作共赢，始终是我们的坚持！
#
# 如果你是有品牌意识，想要定制款式的客户，我们支持原创设计输出！
#
# 🚀 最快的速度，把产品送到你手上！
#
# 无论是产品质量还是服务质量，我们一直在追求更高的标准，致力于越来越好地服务您！
#
# 🛫 一件代发，全球批发，物流可追踪！
#
# 无论您需要少量批发，还是一件代发，我们都能满足您的需求！
#
# ❤️ 多种包装方案，最适合您品牌的包装！
#
# 我们提供多种包装方案，选取最适合您品牌的包装！
# #宠物用品一件代发#宠物用品工厂
#
# 2. 💡 又是为客户发货发到手软的一天啦，打包小哥都快顶不住了呢！我们是宠物用品源头厂家，最近后台收到不少信息，为此统一回复一下大家的疑问。🤔
# 产品定位和优势：我们的产品定位中高端，品类丰富，款式时尚原创，不容易撞款；且一直在更新，高品质售后率低，让你放心选择。👍
# 如何拿货：①一件代发，降低您创业的风险；②批发，起订量低，混批可，减少囤货的压力；③海外市场，帮你一件代发全球，解放你双手和烦恼；④支持定制，每个环节都可以为你安排。客户从小批量拿货到大量自信补货，我们愿意一直和你一起崛起！😍
# 有一个靠谱稳定的进货渠道是非常省心的！如果你也有这个需求欢迎加入我们。💪
# #宠物用品一件代发#宠物用品工厂
#
# 3. 🐾如今养宠物的人数越来越多了！宠物用品市场📈可谓是前途无限！🌠
#
# 👀越来越多的人选择线上购物，大部分是为了更优惠的价格和👌更好的产品质量。我们🏠主营各类猫粮，狗粮，罐头，冻干零食等，货源充足迅捷发货，也有许多支持一件代发的小伙伴们带来了副业收入！🤑
#
# 🤗售后保障服务更是让你没有后顾之忧！☘️爱宠万岁！🐩🐱
# #宠物用品一件代发#宠物用品工厂
#
# 注意，只需要给出一篇正文，且正文里不要出现编号和不要重复标题。"""
# }

class GenerateType(Enum):
    Article = "article"
    Video = "video"


@dataclass_json(undefined="exclude")
@dataclass
class Generation:
    id: str
    type: GenerateType
    titlePrompt: str
    contentPrompt: str
    title: str
    content: str
    urls: list

    createdAt: str
    updatedAt: str


class Generator:
    generate_type: GenerateType

    title_prompt: str
    content_prompt: str

    def __init__(self, generate_type):
        self.generate_type = generate_type

        self.title_prompt = None
        self.content_prompt = None

        self._cache_output = {}

        if "generate" not in initial_state: initial_state["generate"] = {}
        initial_state["generate"][self.name()] = []

        self.check_files()

    def check_files(self):
        # 不存在input目录则创建
        if not os.path.exists(INPUT_ROOT):
            os.makedirs(INPUT_ROOT)
        if not os.path.exists(self.output_dir()):
            os.makedirs(self.output_dir())

        if not os.path.exists(os.path.join(INPUT_ROOT, TITLE_PROMPT_FILE)):
            shutil.copy(
                os.path.join(INPUT_ROOT, "default", TITLE_PROMPT_FILE),
                os.path.join(INPUT_ROOT, TITLE_PROMPT_FILE)
            )

        if not os.path.exists(os.path.join(INPUT_ROOT, CONTENT_PROMPT_FILE)):
            shutil.copy(
                os.path.join(INPUT_ROOT, "default", CONTENT_PROMPT_FILE),
                os.path.join(INPUT_ROOT, CONTENT_PROMPT_FILE)
            )

    def name(self):
        return self.generate_type.value

    # region Config

    def section_name(self):
        return "Generate %s" % self.generate_type.name

    def load_config(self, key, type):
        return get(self.section_name(), key, type)

    def interval(self):
        return self.load_config("interval", "int")

    def max_count(self):
        return self.load_config("max_count", "int")

    # endregion

    # region State

    def generation_ids(self) -> list:
        return get_state("generate", self.name(), default=[])

    def gen_count(self):
        return len(self.generation_ids())

    def _add_count(self, id):
        generation_ids = self.generation_ids()
        generation_ids.append(id)
        set_state(generation_ids, "generate", self.name())

    def _clear_count(self):
        set_state([], "generate", self.name())

    # endregion

    # region Input

    def get_prompts(self) -> (str, str):
        if self.title_prompt is None:
            with open(os.path.join(INPUT_ROOT, TITLE_PROMPT_FILE), "r", encoding="utf8") as file:
                self.title_prompt = file.read()
        if self.content_prompt is None:
            with open(os.path.join(INPUT_ROOT, CONTENT_PROMPT_FILE), "r", encoding="utf8") as file:
                self.content_prompt = file.read()

        return self.title_prompt, self.content_prompt

    # endregion

    # region Output

    def output_dir(self):
        return os.path.join(OUTPUT_ROOT, self.name())

    _cache_output = {}
    def get_output(self, id) -> Generation | None:
        if id in self._cache_output: return self._cache_output[id]

        file_name = os.path.join(self.output_dir(), 'output.xlsx')
        if not os.path.exists(file_name): return None

        self._cache_output[id] = None

        # 获取excel中编号为id的一行数据
        for row in openpyxl.load_workbook(file_name).active.rows:
            if row[0].value == id:
                # 分开获取url转为列表
                temp_urls = [
                    row[i].value for i in range(3, len(row) - 4) if row[i].value != "-"
                ]
                # for i in range(3, len(row) - 4):
                #     if row[i].value != "-":
                #         temp_urls.append(row[i].value)

                self._cache_output[id] = Generation(
                    id=row[0].value,
                    type=self.generate_type,
                    title=row[1].value,
                    content=row[2].value,
                    urls=temp_urls,
                    titlePrompt=row[-4].value,
                    contentPrompt=row[-3].value,
                    createdAt=row[-2].value,
                    updatedAt=row[-1].value
                )
                break
        # with open(file_name, "r", encoding="utf8") as file:

        return self._cache_output[id]

    def _save_generation(self, generation: Generation):
        file_name = os.path.join(self.output_dir(), "output.xlsx")

        # 如果目录不存在，则创建
        # if not os.path.exists(self.output_dir()):
        #     os.makedirs(self.output_dir())

        # 如果文件不存在，则创建
        is_new_file = not os.path.exists(file_name)
        if is_new_file: open(file_name, "w", encoding="utf8").close()

        try:
            workbook = openpyxl.Workbook() if is_new_file else openpyxl.load_workbook(file_name)
            worksheet = workbook.active

            link_font = Font(color="0000FF", underline="single")

            generation_headers = (['ID', '标题', '内容'] +
                       ["素材 #%d" % (i + 1) for i in range(len(generation.urls))] +
                       ['标题提示词', '内容提示词', '创建时间', '更新时间'])
            generation_col = len(generation_headers)

            max_col = worksheet.max_column
            # 确定最大col数量
            real_max_col = max([max_col, generation_col])
            real_url_col = real_max_col - 7

            if is_new_file:
                worksheet.append(generation_headers)
            else:
                # 如果需要，更新标题
                if max_col < real_max_col:
                    # 更新工作表标题
                    for col_num, header_value in enumerate(generation_headers, 1):
                        worksheet.cell(row=1, column=col_num, value=header_value)

                # 更新现有的行以确保每行都有相同数量的列
                for row_num in range(2, worksheet.max_row + 1):
                    curr_row_len = len([cell for cell in worksheet[row_num] if cell.value is not None])
                    if real_max_col == curr_row_len: continue

                    for i in range(4):
                        last_val = worksheet.cell(row=row_num, column=curr_row_len - i).value
                        worksheet.cell(row=row_num, column=real_max_col - i, value=last_val)

                    for i in range(real_max_col - curr_row_len):
                        worksheet.cell(row=row_num, column=curr_row_len - 3 + i, value="-")
                # for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=False):
                #     while len(row) < header_len:
                #         worksheet.cell(row=row[0].row, column=len(row) + 1, value="")

            # 添加新数据
            new_row = ([generation.id, generation.title, generation.content] +
                       generation.urls + ["-"] * (real_url_col - len(generation.urls)) +
                       [generation.titlePrompt, generation.contentPrompt, generation.createdAt, generation.updatedAt])
            worksheet.append(new_row)

            # 创建垂直居中的对齐方式
            alignment = Alignment(wrapText=True, vertical='center')
            # 遍历工作表的所有单元格并设置垂直居中
            for row in worksheet.iter_rows():
                for cell in row: cell.alignment = alignment

            new_row_cells = worksheet[worksheet.max_row]

            # new_row_cells[1].alignment = Alignment(wrapText=True)
            # new_row_cells[2].alignment = Alignment(wrapText=True)
            # new_row_cells[real_max_col - 3].alignment = Alignment(wrapText=True)
            # new_row_cells[real_max_col - 4].alignment = Alignment(wrapText=True)

            for i, url in enumerate(generation.urls, 3):
                if url:
                    new_row_cells[i].hyperlink = os.path.abspath(url)
                    new_row_cells[i].font = link_font

            real_headers = [cell.value for cell in worksheet[1]]
            col_widths = [COL_WIDTH[key.split(" ")[0]] for key in real_headers]

            for (i, col) in enumerate(worksheet.columns):
                column = col[0].column_letter
                worksheet.column_dimensions[column].width = col_widths[i]

            # 保存工作簿
            workbook.save(file_name)
            print("Excel文件保存成功！")

        except Exception as e:
            print("保存Excel文件时出现异常：", str(e))

    # endregion

    # region Generate

    generating_count: int

    def generate(self):
        if self.gen_count() >= self.max_count(): return True

        self.generating_count = self.gen_count() + 1

        print("Start generate %s: %d/%d" % (
            self.name(), self.generating_count, self.max_count()
        ))

        title_prompt, content_prompt, title, content = self._generate_title_content()
        urls = self._generate_media()

        generation = self._upload_generation(title_prompt, content_prompt, title, content, urls)

        self._save_generation(generation)
        self._add_count(generation.id)

        print("End generate %s: %d/%d -> %s" % (
            self.name(), self.generating_count, self.max_count(), generation
        ))

    def _generate_title_content(self):
        if os.path.exists(os.path.join(INPUT_ROOT, TEMPLATE_FILE)):
            title, content = self._get_template(self.generating_count - 1)
            if title is not None and content is not None:
                return "[CUSTOM]", "[CUSTOM]", title, content

        title_prompt, content_prompt = self.get_prompts()
        title = openai_utils.generate_completion(title_prompt)

        content_prompt_with_title = content_prompt.replace("%s", title)
        content = openai_utils.generate_completion(content_prompt_with_title)

        return title_prompt, content_prompt, title, content

    def _get_template(self, index):
        wb = openpyxl.load_workbook(os.path.join(INPUT_ROOT, TEMPLATE_FILE))
        rows = wb.active.rows
        index %= wb.active.max_row - 1
        # index += 1 # 考虑到标题行，所以需要+1

        type = self.generate_type.value.lower()

        for row in rows:
            if row[2].value.lower() != type: continue
            if index == 0: return row[0].value, row[1].value
            index -= 1

        return None, None

    @abstractmethod
    def _generate_media(self) -> list:
        pass

    def _upload_generation(self, title_prompt, content_prompt, title, content, urls) -> Generation:
        return Generation(**api_utils.generate({
            "type": self.generate_type.value,
            "titlePrompt": title_prompt,
            "contentPrompt": content_prompt,
            "title": title,
            "content": content,
            "urls": urls
        }))

    def multi_generate(self):
        while self.gen_count() < self.max_count():
            self.generate()
            time.sleep(self.interval())

    def clear(self):
        if os.path.exists(self.output_dir()):
            shutil.rmtree(self.output_dir())
            os.mkdir(self.output_dir())

        self._clear_count()
        self._clear_publisher()

    def _clear_publisher(self):
        from src.publish.index import PUBLISHERS

        for key in PUBLISHERS:
            if key.endswith(self.name()):
                PUBLISHERS[key].clear()

    # endregion
